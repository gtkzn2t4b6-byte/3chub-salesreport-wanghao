# -*- coding: utf-8 -*-
"""
生成 2026-9 TARGET.xlsx（季节系数 + 单店趋势调整）

公式:
  基础任务 = 近6个月(2026-03~08)平均月销量 × 9月季节系数(0.8837)
  趋势调整(防螺旋下降):
    trend = 近3月均(26-06~08) / 前3月均(26-03~05)
    下滑店(trend<1): 任务 = 基础 × min(1+(1-trend)*0.5, 1.15)   # 补偿一半下滑,封顶+15%
    上升店(trend>=1): 任务 = 基础 × min(1+(trend-1)*0.25, 1.10) # 小幅上调,封顶+10%
  历史缺失门店: 8月任务 × (0.8837/1.0045) 季节等比换算, 不做趋势调整

输出(在8月文件副本上原位改值, 保留全部格式):
  ~/Desktop/销售部/增量基数/2026年/2026-9 TARGET.xlsx
  sheet: 任务汇总 / 手机 / 配件(不变)
"""
import json
import shutil
import numpy as np
import pandas as pd
from openpyxl import load_workbook

SEP_COEF = 0.8837     # 9月季节系数 (2022年季节系数表)
AUG_COEF = 1.0045     # 8月季节系数
FALL_COMP, FALL_CAP = 0.5, 0.15   # 下滑补偿比例/封顶
RISE_COMP, RISE_CAP = 0.25, 0.10  # 上升上调比例/封顶

# 已关店门店(2026-09起不再下任务, 从目标文件删除):
CLOSED_STORES = {
    'D_IGANDO-IKOTUN ROAD-LAGOS',        # IGANDO
    'D_MSL MUSHIN2-ISOLO ROAD-LAGOS',    # MSL MUSHIN2
}

SRC = '/Users/wanghao/Desktop/销售部/增量基数/2026年/2026-8 TARGET.xlsx'
OUT = '/Users/wanghao/Desktop/销售部/增量基数/2026年/2026-9 TARGET.xlsx'
HIST = '/Users/wanghao/WorkBuddy/2026-07-02-11-18-30/historical_data.json'

MONTHS6 = [f'2026-{m:02d}' for m in range(3, 9)]   # 26-03..08
RECENT3 = [f'2026-{m:02d}' for m in range(6, 9)]   # 26-06..08
PREV3 = [f'2026-{m:02d}' for m in range(3, 6)]     # 26-03..05

# ---------- 1. 读取历史与8月目标 ----------
hist = json.load(open(HIST))['data']['phone_sales']  # store(无D_前缀) -> {'YYYY-MM': qty}

aug_tgt = pd.read_excel(SRC, sheet_name='手机')
aug_tgt = aug_tgt[aug_tgt['门店'] != '合计'].copy()
aug_map = dict(zip(aug_tgt['门店'], aug_tgt['任务']))  # D_xxx -> 8月任务

# ---------- 2. 计算各店9月任务 ----------
def mean(vals):
    vals = [v for v in vals if v is not None]
    return float(np.mean(vals)) if vals else None

records = []
for store, aug_t in aug_map.items():
    if store in CLOSED_STORES:      # 关店门店不再计算任务
        continue
    key = store[2:] if store.startswith('D_') else store
    m = hist.get(key)
    if m is None:
        records.append(dict(store=store, aug=aug_t, a6=None, trend=None,
                            base=aug_t * SEP_COEF / AUG_COEF, factor=1.0,
                            task=aug_t * SEP_COEF / AUG_COEF, note='无历史,季节换算'))
        continue
    a6 = mean([m.get(x) for x in MONTHS6])
    rec = mean([m.get(x) for x in RECENT3])
    pre = mean([m.get(x) for x in PREV3])
    if a6 is None:
        records.append(dict(store=store, aug=aug_t, a6=None, trend=None,
                            base=aug_t * SEP_COEF / AUG_COEF, factor=1.0,
                            task=aug_t * SEP_COEF / AUG_COEF, note='无近期数据'))
        continue
    trend = rec / pre if (pre and pre > 0 and rec is not None) else 1.0
    base = a6 * SEP_COEF
    if trend < 1:
        factor = min(1 + (1 - trend) * FALL_COMP, 1 + FALL_CAP)
        note = '下滑店,补偿一半'
    else:
        factor = min(1 + (trend - 1) * RISE_COMP, 1 + RISE_CAP)
        note = '上升店,小幅上调' if factor > 1 else '平稳'
    records.append(dict(store=store, aug=aug_t, a6=round(a6, 1), trend=round(trend, 3),
                        base=base, factor=round(factor, 3), task=base * factor, note=note))

df = pd.DataFrame(records)
# 任务取整: 向上取整到10的倍数(如289.2→290), 不留小数, 个位数任务自动补到10
df['task'] = (np.ceil(df['task'] / 10) * 10).astype(int)
new_map = dict(zip(df['store'], df['task']))

# ---------- 3. 原位改值生成新文件 ----------
# 注意: 8月文件「任务汇总」大量使用公式(含外部文件引用 VLOOKUP [1]完成度),
# openpyxl 往返会丢失公式缓存值导致 pandas 读到 NaN, 因此所有公式一律替换为字面值
aug_names = pd.read_excel(SRC, sheet_name='任务汇总', header=None, skiprows=2)
aug_names.columns = ['x0', 'shop', 'name', 'phone', 'inst']
aug_names = aug_names[aug_names['shop'].astype(str).str.startswith('D_')]
name_map = dict(zip(aug_names['shop'].astype(str).str.strip(), aug_names['name']))

shutil.copy(SRC, OUT)
wb = load_workbook(OUT)

# 手机 sheet: col A=门店, B=任务; 首个数据行(第2行)为合计
ws = wb['手机']
total = int(df['task'].sum())
ws.cell(row=2, column=2).value = total  # 合计行(各店均为10倍数,合计自动为10倍数整数)
n_phone = 0
for r in range(3, ws.max_row + 1):
    s = ws.cell(row=r, column=1).value
    if s and str(s).strip() in new_map:
        ws.cell(row=r, column=2).value = new_map[str(s).strip()]
        n_phone += 1

# 任务汇总: col B=SHOP NAME, C=NAME, D=手机任务, E=分期任务
# 注意: 门店并非严格按 3C段→MSL段 顺序排列(存在3C门店夹在MSL群中, 如D_TRS APAPA),
#       因此必须分两遍: 先写完所有门店并累加, 再统一写小计/合计行, 否则小计会漏后续门店
ws2 = wb['任务汇总']
sub3c = submsl = grand = 0.0
n_sum = 0
# 第一遍: 只写门店行, 累加小计
for r in range(3, ws2.max_row + 1):
    shop = ws2.cell(row=r, column=2).value
    if shop and str(shop).strip() in new_map:
        s = str(shop).strip()
        v = new_map[s]
        ws2.cell(row=r, column=3).value = name_map.get(s)  # NAME 公式→字面值
        ws2.cell(row=r, column=4).value = int(v)
        # 分期任务 = 月任务×30%(月任务为10倍数→分期为整数, 如1450→435, 不强制10倍数)
        ws2.cell(row=r, column=5).value = int(round(v * 0.3))
        if s.startswith('D_MSL'):
            submsl += v
        else:
            sub3c += v
        grand += v
        n_sum += 1
# 第二遍: 所有门店值累加完毕后再写小计/合计行
for r in range(3, ws2.max_row + 1):
    shop = ws2.cell(row=r, column=2).value
    name = ws2.cell(row=r, column=3).value
    if name == '3C HUB':
        ws2.cell(row=r, column=4).value = int(round(sub3c))
        ws2.cell(row=r, column=5).value = int(round(sub3c * 0.3))
    elif name == 'MSL' and not shop:
        ws2.cell(row=r, column=4).value = int(round(submsl))
        ws2.cell(row=r, column=5).value = int(round(submsl * 0.3))
    elif shop == '(全部)':
        ws2.cell(row=r, column=4).value = int(round(grand))
        ws2.cell(row=r, column=5).value = int(round(grand * 0.3))

# 删除已关店门店所在行(从后往前删除,避免行号错位)
del_ph = [r for r in range(3, ws.max_row + 1)
          if ws.cell(row=r, column=1).value
          and str(ws.cell(row=r, column=1).value).strip() in CLOSED_STORES]
for r in sorted(del_ph, reverse=True):
    ws.delete_rows(r)
del_sm = [r for r in range(3, ws2.max_row + 1)
          if ws2.cell(row=r, column=2).value
          and str(ws2.cell(row=r, column=2).value).strip() in CLOSED_STORES]
for r in sorted(del_sm, reverse=True):
    ws2.delete_rows(r)
assert del_ph and len(del_ph) == len(CLOSED_STORES), f'手机sheet关店行删除异常: {del_ph}'
assert del_sm and len(del_sm) == len(CLOSED_STORES), f'任务汇总关店行删除异常: {del_sm}'
print(f'已删除关店门店行 {len(del_ph)} 家: 手机sheet+任务汇总 (IGANDO / MSL MUSHIN2)')

# 兜底: 确认任务汇总无残留公式
leftover = [c.coordinate for row in ws2.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith('=')]
assert not leftover, f'任务汇总残留公式: {leftover}'
leftover2 = [c.coordinate for row in ws.iter_rows() for c in row
             if isinstance(c.value, str) and c.value.startswith('=')]
assert not leftover2, f'手机sheet残留公式: {leftover2}'

wb.save(OUT)

# ---------- 4. 汇总输出 ----------
n_total = len(df)
print(f'手机sheet更新门店: {n_phone}/{n_total}, 任务汇总更新: {n_sum}/{n_total}')
print(f'9月总目标: {total:.0f} 台  (含{len(CLOSED_STORES)}家关店门店已剔除, 8月目标 36,923, 8月实际 32,034)')
print(f'3C HUB小计: {sub3c:.0f}  MSL小计: {submsl:.0f}')
print()
print('== 趋势调整分布 ==')
print(df['note'].value_counts().to_string())
print()
print('== 与8月目标对比(变化最大) ==')
df['chg'] = (df['task'] / df['aug'] - 1) * 100
cols = ['store', 'aug', 'a6', 'trend', 'factor', 'task', 'chg']
print(df.nlargest(5, 'chg')[cols].round(1).to_string(index=False, max_colwidth=32))
print(df.nsmallest(5, 'chg')[cols].round(1).to_string(index=False, max_colwidth=32))
print()
print('== Top15门店 ==')
print(df.nlargest(15, 'task')[cols].round(1).to_string(index=False, max_colwidth=32))
df.to_csv('/Users/wanghao/WorkBuddy/2026-07-02-11-18-30/scripts/target_2026_9_detail.csv', index=False)
print('\n明细已存 scripts/target_2026_9_detail.csv')
